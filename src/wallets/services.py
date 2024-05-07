import asyncio
from typing import Any, Dict, List, Sequence, Type

import aiohttp
from database import async_session
from eth_typing import Address, ChecksumAddress
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from unit_of_work import UnitOfWork
from wallets.config import ABI, CHAINS
from wallets.models import Wallet, WalletGroup
from wallets.schemas import (
    ChainBalanceSchema,
    ChainSchema,
    WalletBalanceSchema,
    WalletCreateSchema,
    WalletDeleteSchema,
    WalletGroupCreateSchema,
    WalletGroupPatchSchema,
    WalletGroupPutSchema,
    WalletPatchSchema,
    WalletPutSchema,
)
from web3 import AsyncHTTPProvider, AsyncWeb3


class WalletService:
    def __init__(self, unit_of_work: Type[UnitOfWork]):
        self._unit_of_work = unit_of_work(async_session)

    async def get_wallets(self, user_id: int) -> Sequence[Wallet]:
        async with self._unit_of_work as uow:
            wallets = await uow.wallet.get_multiple_by(user_id=user_id)

            return wallets

    async def get_wallet(self, wallet_id: int, user_id: int) -> Wallet:
        async with self._unit_of_work as uow:
            wallet = await uow.wallet.get_by(id=wallet_id)

            if not wallet:
                raise HTTPException(status_code=404, detail="Wallet not found")

            if wallet.user_id != user_id:
                raise HTTPException(
                    status_code=403, detail="Wallet belongs to another user"
                )

            return wallet

    async def update_wallets(
        self, wallets: List[WalletPatchSchema] | List[WalletPutSchema], user_id: int
    ) -> Sequence[Wallet]:
        async with self._unit_of_work as uow:
            wallet_ids: List[int] = [wallet.id for wallet in wallets]

            wallets_to_update = await self._get_and_validate_wallets(
                uow, wallet_ids, user_id
            )

            wallet_mappings = [
                wallet.model_dump(exclude_none=True) for wallet in wallets
            ]

            updated_wallets = await uow.wallet.update_multiple_wallets(
                wallets_to_update, wallet_mappings
            )

            await uow.commit()

            return updated_wallets

    async def delete_wallets(
        self, wallet_ids: List[WalletDeleteSchema], user_id: int
    ) -> None:
        async with self._unit_of_work as uow:
            wallet_ids_to_delete = [wallet.id for wallet in wallet_ids]

            await self._get_and_validate_wallets(uow, wallet_ids_to_delete, user_id)

            await uow.wallet.delete_multiple_wallets(wallet_ids_to_delete, user_id)

            await uow.commit()

    async def _get_and_validate_wallets(
        self, uow: UnitOfWork, wallet_ids: List[int], user_id: int
    ) -> Sequence[Wallet]:
        wallets = await uow.wallet.filter_by_wallet_ids(wallet_ids)

        if len(wallets) != len(wallet_ids):
            raise HTTPException(status_code=400, detail="Invalid wallet IDs")

        for wallet in wallets:
            if wallet.user_id != user_id:
                raise HTTPException(
                    status_code=403,
                    detail="Some of the wallets belongs to another user",
                )

        return wallets


class WalletImporterService:
    def __init__(self, unit_of_work: Type[UnitOfWork]) -> None:
        self._unit_of_work = unit_of_work(async_session)

    async def import_wallets(
        self, wallets: List[WalletCreateSchema], user_id: int
    ) -> Sequence[Wallet]:
        async with self._unit_of_work as uow:
            existing_wallets = await uow.wallet.get_multiple_by(user_id=user_id)
            existing_addresses = {wallet.address for wallet in existing_wallets}
            max_numbered_wallet = self._get_max_numbered_wallet(existing_wallets)

            wallets_data: List[Dict[str, Any]] = []
            for wallet in wallets:
                if wallet.address in existing_addresses:
                    continue

                max_numbered_wallet += 1
                wallet_data = {
                    "number": max_numbered_wallet,
                    "address": wallet.address,
                    "user_id": user_id,
                }
                wallets_data.append(wallet_data)

            imported_wallets = await uow.wallet.create_multiple(wallets_data)

            await uow.commit()

            return imported_wallets

    async def import_wallets_xlsx(
        self, wallets_xlsx: UploadFile, user_id: int
    ) -> Sequence[Wallet]:
        async with self._unit_of_work as uow:
            wb = load_workbook(filename=wallets_xlsx.file)
            sheet = wb.active

            existing_wallets = await uow.wallet.get_multiple_by(user_id=user_id)
            existing_addresses = {wallet.address for wallet in existing_wallets}
            max_numbered_wallet = self._get_max_numbered_wallet(existing_wallets)

            wallets_data: List[Dict[str, Any]] = []

            for row in sheet.iter_rows(min_rows=1):
                if row[0].value in existing_addresses:
                    continue

                max_numbered_wallet += 1
                wallet_data = {
                    "number": max_numbered_wallet,
                    "address": row[0].value,
                    "user_id": user_id,
                }
                wallets_data.append(wallet_data)

            imported_wallets = await uow.wallet.create_multiple(wallets_data)

            await uow.commit()

            return imported_wallets

    def _get_max_numbered_wallet(self, existing_wallets: Sequence[Wallet]) -> int:
        numbered_wallets = [wallet.number for wallet in existing_wallets]

        return max(numbered_wallets, default=0)


class WalletGroupService:
    def __init__(self, unit_of_work: Type[UnitOfWork]) -> None:
        self._unit_of_work = unit_of_work(async_session)

    async def get_wallet_groups(self, user_id: int) -> Sequence[WalletGroup]:
        async with self._unit_of_work as uow:
            return await uow.wallet_group.get_multiple_by(user_id=user_id)

    async def get_wallet_group(self, wallet_group_id: int, user_id: int) -> WalletGroup:
        async with self._unit_of_work as uow:
            wallet_group = await uow.wallet_group.get_by(id=wallet_group_id)

            if not wallet_group:
                raise HTTPException(status_code=404, detail="Wallet group not found")

            if wallet_group.user_id != user_id:
                raise HTTPException(
                    status_code=403, detail="Wallet group belongs to another user"
                )

            return wallet_group

    async def create_wallet_group(
        self, wallet_group: WalletGroupCreateSchema, user_id: int
    ) -> WalletGroup:
        async with self._unit_of_work as uow:
            wallet_group_data = wallet_group.model_dump(exclude_none=True)
            wallet_group_data["user_id"] = user_id

            if await uow.wallet_group.get_by(name=wallet_group_data["name"]):
                raise HTTPException(
                    status_code=409, detail="Wallet group with this name already exists"
                )

            created_wallet_group = await uow.wallet_group.create(wallet_group_data)

            await uow.commit()

            return created_wallet_group

    async def update_wallet_group(
        self,
        wallet_group_id: int,
        wallet_group: WalletGroupPatchSchema | WalletGroupPutSchema,
        user_id: int,
    ) -> WalletGroup:
        async with self._unit_of_work as uow:
            wallet_group_to_update = await uow.wallet_group.get_by(id=wallet_group_id)

            if not wallet_group_to_update:
                raise HTTPException(status_code=404, detail="Wallet group not found")

            if wallet_group_to_update.user_id != user_id:
                raise HTTPException(
                    status_code=403, detail="You can only update your wallet groups"
                )

            wallet_group_data = wallet_group.model_dump(exclude_none=True)

            updated_wallet_group = await uow.wallet_group.update(
                wallet_group_to_update, wallet_group_data
            )

            await uow.commit()

            return updated_wallet_group

    async def delete_wallet_group(self, wallet_group_id: int, user_id: int) -> None:
        async with self._unit_of_work as uow:
            wallet_group_to_delete = await uow.wallet_group.get_by(id=wallet_group_id)

            if not wallet_group_to_delete:
                raise HTTPException(status_code=404, detail="Wallet group not found")

            if wallet_group_to_delete.user_id != user_id:
                raise HTTPException(
                    status_code=403, detail="You can only delete your wallet groups"
                )

            await uow.wallet_group.delete(wallet_group_to_delete)

            await uow.commit()


class BalanceService:
    def __init__(self, unit_of_work: Type[UnitOfWork]) -> None:
        self._unit_of_work = unit_of_work(async_session)

    async def get_wallets_balance(
        self, user_id: int, selected_chains: List[ChainSchema]
    ) -> List[ChainBalanceSchema]:
        tasks = list()
        async with self._unit_of_work as uow:
            wallets = await uow.wallet.get_multiple_by(user_id=user_id)

            for wallet in wallets:
                wallet_address = wallet.address

                for selected_chain in selected_chains:
                    chain = selected_chain.name
                    chain_info = CHAINS[selected_chain.name]
                    tasks.append(self._process_chain(wallet_address, chain, chain_info))

            results = await asyncio.gather(*tasks)

            wallet_balances = [
                ChainBalanceSchema(**result) for result in results if result
            ]

            return wallet_balances

    async def _process_chain(
        self, wallet_address: str, chain: str, chain_info
    ) -> dict[str, Any] | None:
        for rpc in chain_info["rpc"]:
            try:
                web3 = AsyncWeb3(AsyncHTTPProvider(rpc))

                wallet_address = web3.to_checksum_address(wallet_address)

                if await web3.eth.get_transaction_count(wallet_address) == 0:
                    return

                native_balance = await self._get_native_balance(web3, wallet_address)

                native_in_usd = await self._native_balance_in_usd(
                    chain_info["currency"], native_balance
                )

                usdt_balance = None
                if "usdt_contract_address" in chain_info:
                    usdt_contract_address = chain_info["usdt_contract_address"]
                    usdt_balance = await self._get_contract_balance(
                        web3, wallet_address, usdt_contract_address
                    )

                usdc_balance = None
                if "usdc_contract_address" in chain_info:
                    usdc_contract_address = chain_info["usdc_contract_address"]
                    usdc_balance = await self._get_contract_balance(
                        web3, wallet_address, usdc_contract_address
                    )

                return {
                    "chain": chain,
                    "balance": WalletBalanceSchema(
                        address=wallet_address,
                        native_balance=native_balance,
                        native_in_usd=native_in_usd,
                        usdt_balance=usdt_balance,
                        usdc_balance=usdc_balance,
                    ),
                }

            except aiohttp.ClientResponseError:
                continue

    async def _get_native_balance(
        self, web3: AsyncWeb3, wallet_address: ChecksumAddress
    ) -> int | float:
        native_balance_wei = await web3.eth.get_balance(wallet_address)
        native_balance = web3.from_wei(native_balance_wei, "ether")

        if isinstance(native_balance, int):
            return native_balance

        return float(f"{native_balance:.6f}")

    async def _native_balance_in_usd(
        self, currency: str, native_balance: int | float
    ) -> int | float:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                f"https://min-api.cryptocompare.com/data/price?fsym={currency}&tsyms=USD"
            ) as resp:
                if resp.status != 200:
                    raise HTTPException(
                        status_code=500, detail="Failed to get USD price"
                    )

                usd_price = (await resp.json())["USD"]

                return float(f"{native_balance * usd_price:.6f}")

    async def _get_contract_balance(
        self,
        web3: AsyncWeb3,
        wallet_address: ChecksumAddress,
        usdt_contract_address: Address,
    ) -> int | float:
        usdt_contract = web3.eth.contract(address=usdt_contract_address, abi=ABI)

        contract_balance = await usdt_contract.functions.balanceOf(
            wallet_address
        ).call()
        decimals = await usdt_contract.functions.decimals().call()

        contract_balance = contract_balance / 10**decimals

        if contract_balance.is_integer():
            return int(contract_balance)

        return float(f"{contract_balance:.6f}")
